import os
import re
import json
import base64
import tempfile
from io import BytesIO

from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

load_dotenv()

app = Flask(__name__, static_folder="public", static_url_path="")
CORS(app)

# Configure Groq via the OpenAI-compatible API
API_KEY = os.getenv("GROQ_API_KEY")
if not API_KEY:
    raise RuntimeError("GROQ_API_KEY not found in .env file")

client = OpenAI(
    api_key=API_KEY,
    base_url="https://api.groq.com/openai/v1",
)

# Default to Groq-hosted Llama 4 Scout, a vision-capable model with broader availability.
MODEL = os.getenv("GROQ_MODEL", "meta-llama/llama-4-scout-17b-16e-instruct")
VISION_MODELS = {
    "meta-llama/llama-4-scout-17b-16e-instruct",
    "meta-llama/llama-4-maverick-17b-128e-instruct",
}

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "bmp", "gif"}
MIME_MAP = {
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "webp": "image/webp",
    "bmp": "image/bmp",
    "gif": "image/gif",
}

# ─── Extraction Prompt ────────────────────────────────────────────────
# Each line is carefully crafted for financial report extraction.
# The prompt is structured in 4 sections:
#   1. ROLE      – sets the LLM's expertise context
#   2. TASK      – what exactly to extract
#   3. RULES     – how to handle financial report quirks
#   4. OUTPUT    – strict JSON format specification

EXTRACTION_PROMPT = """
You are a financial data extraction specialist.
Your task is to extract ALL tabular data from
the provided financial report screenshot.

This image is from a financial statement such as:
- Balance Sheet (Statement of Financial Position)
- Profit & Loss / Income Statement
- Cash Flow Statement
- Schedule of Notes to Accounts
- Statement of Changes in Equity

─── EXTRACTION RULES ───

1. EVERY row in the table must be captured — do NOT skip any line items,
   including section headers, subtotals, totals, and blank separator rows.

2. Preserve the EXACT hierarchy of the financial statement.
   Add a "Level" field to indicate depth:
   - "header"  → section headings (e.g. "Non-Current Assets", "Equity and Liabilities")
   - "item"    → individual line items (e.g. "Property, Plant and Equipment")
   - "subtotal"→ sub-totals   (e.g. "Total Non-Current Assets")
   - "total"   → grand totals (e.g. "Total Assets", "Total Equity and Liabilities")

3. For the "Note" or "Notes" column:
   - Keep the note reference number exactly as shown (e.g. "1", "2a", "III")
   - If a row has no note reference, use an empty string ""

4. NUMERIC VALUES — critical rules:
   - Remove any comma separators (e.g. "2,67,096" → "267096")
   - Keep negative numbers as negatives: "(1,234)" → "-1234"
   - Parenthesized numbers in financial reports are NEGATIVE values
   - Keep decimals if present (e.g. "12.50" stays "12.50")
   - If a cell shows "—" or "-" or is blank, use ""
   - Do NOT fabricate or estimate any numbers

5. COLUMN HEADERS:
   - Use the EXACT column headers as they appear in the document
   - For date-based columns, keep the full date (e.g. "As at 31st March, 2025")
   - Multi-line headers should be merged into one string

6. If there are MULTIPLE tables in the image (e.g. Assets on top, Equity & Liabilities below),
   combine them into a SINGLE flat array. Use the "Section" field to distinguish which
   part of the statement each row belongs to (e.g. "Assets", "Equity and Liabilities").

7. Currency symbols (₹, $, €, etc.) should NOT appear in numeric values.
   Mention the currency unit in a "Currency" field only if shown in the header.

─── OUTPUT FORMAT ───

Return ONLY a valid JSON array. No markdown, no explanation, no commentary.

Each object in the array must have EXACTLY these keys (plus any date columns):
{
  "Section": "which part of the financial statement",
  "Level": "header | item | subtotal | total",
  "Line Item": "the row label exactly as shown",
  "Notes": "note reference or empty string",
  "<Period Column 1>": "numeric value as string",
  "<Period Column 2>": "numeric value as string"
}

Where <Period Column 1>, <Period Column 2>, etc. are the actual date/period
column headers from the image (e.g. "As at 31st March, 2025").

CRITICAL: Output ONLY the JSON array. No text before or after it.
"""

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def get_mime(filename):
    ext = filename.rsplit(".", 1)[1].lower()
    return MIME_MAP.get(ext, "image/png")


def extract_table_from_image(image_path, mime_type):
    """Send image to Groq and extract table data as JSON."""
    with open(image_path, "rb") as f:
        image_data = base64.b64encode(f.read()).decode("utf-8")

    if MODEL not in VISION_MODELS:
        raise RuntimeError(
            f"Configured model '{MODEL}' does not support image input on Groq. "
            "Set GROQ_MODEL to a vision-capable Groq model for screenshot extraction."
        )

    response = client.responses.create(
        model=MODEL,
        input=[
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": EXTRACTION_PROMPT},
                    {
                        "type": "input_image",
                        "detail": "auto",
                        "image_url": f"data:{mime_type};base64,{image_data}",
                    },
                ],
            }
        ],
        temperature=1e-8,
        max_output_tokens=8192,
    )

    text = response.output_text
    print(f"--- RAW GROQ RESPONSE ---\n{text}\n--- END ---")

    return parse_json_from_text(text)


def parse_json_from_text(text):
    """Robustly extract a JSON array from model output that may contain extra text."""
    # Strip markdown code fences
    text = re.sub(r"```(?:json)?\s*", "", text).strip()

    # Try parsing with multiple strategies
    for candidate in _extract_candidates(text):
        try:
            data = json.loads(candidate)
            if isinstance(data, list):
                return data
            if isinstance(data, dict):
                return [data]
        except json.JSONDecodeError:
            pass

        # Try repairing common model JSON errors before giving up on this candidate
        repaired = _repair_json(candidate)
        try:
            data = json.loads(repaired)
            if isinstance(data, list):
                return data
            if isinstance(data, dict):
                return [data]
        except json.JSONDecodeError:
            pass

    raise json.JSONDecodeError("Could not find valid JSON in model response", text, 0)


def _extract_candidates(text):
    """Yield JSON candidate strings from text, from most to least specific."""
    # 1) Full text
    yield text

    # 2) Outermost JSON array [ ... ]
    match = re.search(r"\[[\s\S]*\]", text)
    if match:
        yield match.group()

    # 3) Outermost JSON object { ... }
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        yield match.group()


def _repair_json(text):
    """Fix common JSON errors produced by LLMs.

    Handles patterns like:
      "Total Assets": ""   →  used as a key, producing  "key1": "key2": "val"
    which is invalid JSON. We fix by collapsing double-colon key patterns.
    """
    # Fix  "key": "value": "value2"  →  "key": "value2"
    # This happens when the model puts a sub-header as a key-value pair
    text = re.sub(
        r'"([^"]*)":\s*"([^"]*)":\s*"([^"]*)"',
        r'"\1 \2": "\3"',
        text,
    )
    # Remove trailing commas before } or ]
    text = re.sub(r",\s*([}\]])", r"\1", text)
    return text


def create_excel(data):
    """Convert a list of dicts into a styled Excel workbook.

    Formatting matches Sample.xlsx:
    - Calibri 11pt throughout
    - Dark blue (#1F4E78) column headers with white bold text
    - Light blue (#DDEBF7) section header rows with bold text
    - Bold + top thin border for subtotals/totals
    - Indent = 1 for regular line items
    - Number format: #,##0;(#,##0);"-" for numeric columns
    - Numeric values colored dark blue (#1F4E78)
    - No cell borders on regular data rows
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    if not data:
        return wb

    # Determine which columns are in the data
    all_keys = list(data[0].keys())

    # Separate metadata keys from value (period) columns
    meta_keys = []
    period_keys = []
    for key in all_keys:
        if key in ("Section", "Level", "Line Item", "Notes"):
            meta_keys.append(key)
        else:
            period_keys.append(key)

    # We only write: Line Item, Notes, then period columns
    # (Section and Level are used for formatting but not shown as columns)
    display_cols = []
    if "Line Item" in meta_keys:
        display_cols.append("Line Item")
    if "Notes" in meta_keys:
        display_cols.append("Notes")
    display_cols.extend(period_keys)

    # ── Styles ──────────────────────────────────────────────
    # Column header row: dark blue background, white bold text
    col_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    col_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    # Section header rows (e.g. "Balance Sheet", "Profit & Loss"): light blue bg, bold
    section_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    section_font = Font(name="Calibri", size=11, bold=True)

    # Category headers (e.g. "ASSETS", "Non-current Assets"): bold, no fill
    category_font = Font(name="Calibri", size=11, bold=True)

    # Subtotal/total rows: bold text, thin top border
    subtotal_font = Font(name="Calibri", size=11, bold=True)
    top_border = Border(top=Side(style="thin"))

    # Regular data items
    item_font = Font(name="Calibri", size=11)
    item_align_indent = Alignment(horizontal="left", indent=1)
    item_align_normal = Alignment(horizontal="left")

    # Numeric values: dark blue color, custom number format
    num_font = Font(name="Calibri", size=11, color="1F4E78")
    num_font_bold = Font(name="Calibri", size=11, color="1F4E78", bold=True)
    num_format = '#,##0.00;\\(#,##0.00\\);"-"'

    # ── Column header row ───────────────────────────────────
    for col_idx, col_name in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = col_header_font
        cell.fill = col_header_fill

    # ── Data rows ───────────────────────────────────────────
    current_row = 2  # start after header

    for row_data in data:
        level = row_data.get("Level", "item").lower()
        section = row_data.get("Section", "")
        line_item = row_data.get("Line Item", "")

        # -- Write the Line Item cell --
        li_col = 1 if "Line Item" in display_cols else None
        if li_col is not None:
            cell = ws.cell(row=current_row, column=li_col, value=line_item)

            if level == "header":
                # Check if this is a major section header (like "Balance Sheet")
                # or a category header (like "ASSETS", "Non-current Assets")
                cell.font = category_font
            elif level in ("subtotal", "total"):
                cell.font = subtotal_font
                cell.border = top_border
            else:
                # Regular item — indented
                cell.font = item_font
                cell.alignment = item_align_indent

        # -- Write the Notes cell --
        notes_col = display_cols.index("Notes") + 1 if "Notes" in display_cols else None
        if notes_col is not None:
            notes_val = row_data.get("Notes", "")
            cell = ws.cell(row=current_row, column=notes_col, value=notes_val)
            cell.font = item_font
            if level in ("subtotal", "total"):
                cell.border = top_border

        # -- Write period/numeric columns --
        for period_key in period_keys:
            col_idx = display_cols.index(period_key) + 1
            raw_val = row_data.get(period_key, "")

            # Try to convert to a number
            numeric_val = _to_number(raw_val)

            cell = ws.cell(row=current_row, column=col_idx)

            if numeric_val is not None:
                cell.value = numeric_val
                cell.number_format = num_format
                if level in ("subtotal", "total"):
                    cell.font = num_font_bold
                    cell.border = top_border
                else:
                    cell.font = num_font
            else:
                cell.value = raw_val if raw_val else ""
                if level in ("subtotal", "total"):
                    cell.font = subtotal_font
                    cell.border = top_border
                elif level == "header":
                    cell.font = category_font

        # -- Apply section fill for "header" level rows that look like section titles --
        if level == "header" and line_item and any(
            keyword in line_item.lower()
            for keyword in ("balance sheet", "profit", "loss", "cash flow", "statement")
        ):
            for col_idx in range(1, len(display_cols) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.fill = section_fill
                cell.font = section_font

        current_row += 1

    # ── Column widths ───────────────────────────────────────
    for col_idx, col_name in enumerate(display_cols, 1):
        if col_name == "Line Item":
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 45
        elif col_name == "Notes":
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 8
        else:
            # Period columns — fit to header length + padding
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(col_name) + 4, 18)

    return wb


def _to_number(val):
    """Try to convert a string value to a float. Returns None if not numeric."""
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return val
    val = str(val).strip()
    # Handle parenthesized negatives: (1234) → -1234
    if val.startswith("(") and val.endswith(")"):
        val = "-" + val[1:-1]
    # Remove commas
    val = val.replace(",", "")
    # Remove currency symbols
    val = val.replace("₹", "").replace("$", "").replace("€", "").strip()
    try:
        return float(val)
    except ValueError:
        return None


# ─── Routes ───────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory("public", "index.html")


@app.route("/preview", methods=["POST"])
def preview():
    """Upload image → return extracted JSON for in-page preview."""
    if "image" not in request.files:
        return jsonify({"error": "No image file provided"}), 400

    file = request.files["image"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    if not allowed_file(file.filename):
        return jsonify({"error": f"Unsupported file type. Use: {', '.join(ALLOWED_EXTENSIONS)}"}), 400

    # Save temp file
    tmp = tempfile.NamedTemporaryFile(delete=False, dir=UPLOAD_DIR, suffix=os.path.splitext(file.filename)[1])
    file.save(tmp.name)

    try:
        mime = get_mime(file.filename)
        data = extract_table_from_image(tmp.name, mime)
        return jsonify({"data": data})
    except json.JSONDecodeError:
        return jsonify({"error": "Could not parse table data from image. Try a clearer screenshot."}), 422
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        os.unlink(tmp.name)


@app.route("/convert", methods=["POST"])
def convert():
    """Upload image → return Excel file download."""
    if "image" not in request.files:
        return jsonify({"error": "No image file provided"}), 400

    file = request.files["image"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    if not allowed_file(file.filename):
        return jsonify({"error": f"Unsupported file type. Use: {', '.join(ALLOWED_EXTENSIONS)}"}), 400

    tmp = tempfile.NamedTemporaryFile(delete=False, dir=UPLOAD_DIR, suffix=os.path.splitext(file.filename)[1])
    file.save(tmp.name)

    try:
        mime = get_mime(file.filename)
        data = extract_table_from_image(tmp.name, mime)

        wb = create_excel(data)
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="extracted_data.xlsx",
        )
    except json.JSONDecodeError:
        return jsonify({"error": "Could not parse table data from image. Try a clearer screenshot."}), 422
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        os.unlink(tmp.name)


if __name__ == "__main__":
    print("Server running at http://localhost:8000")
    app.run(debug=True, host="::", port=8000)
